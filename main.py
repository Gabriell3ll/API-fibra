#!/usr/bin/env python3
# calcular_rutas_full_complete.py
# Versión completa: lógica de cálculo + API HTTP lista para n8n / curl / Render
# Requisitos (pip): openpyxl flask gunicorn requests
# Uso (desarrollo): python calcular_rutas_full_complete.py
# Uso (producción): gunicorn -w 4 "calcular_rutas_full_complete:app"

from datetime import date
import json, math, shutil, os, sys
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from flask import Flask, request, jsonify, send_from_directory, make_response
from werkzeug.utils import secure_filename

# ====================
# CONFIG (entorno)
# ====================
PLANTILLA = os.environ.get("PLANTILLA", "Plantilla.xlsm")
OUTPUT_PREFIX = os.environ.get("OUTPUT_PREFIX", "Resultado")
OUTPUT_DIR = os.environ.get("OUTPUT_DIR", "/tmp")
START_ROW = int(os.environ.get("START_ROW", 3))
MAX_UPLOAD_BYTES = int(os.environ.get("MAX_UPLOAD_BYTES", 8 * 1024 * 1024))

# Asegurar directorio de salida
os.makedirs(OUTPUT_DIR, exist_ok=True)

# --------------------
# Helpers equivalentes a VBA
# --------------------
def NivelKV(tipoKV: str):
    m = {"KV100":1, "KV200":2, "KV300":3, "KV500":4, "KV1000":5}
    return m.get(str(tipoKV or "").strip().upper(), 0)

def TipoDesdeNivel(nivel:int):
    r = {1:"KV100",2:"KV200",3:"KV300",4:"KV500",5:"KV1000"}
    return r.get(nivel, "")

def distancia_m(lat1, lon1, lat2, lon2):
    try:
        R = 6371000.0
        dx = math.radians(lon2 - lon1) * R * math.cos(math.radians((lat1 + lat2) / 2.0))
        dy = math.radians(lat2 - lat1) * R
        return math.sqrt(dx*dx + dy*dy) * 1.05
    except Exception:
        return None

def angle_deg(lat1, lon1, lat2, lon2, lat3, lon3):
    try:
        v1x = (lon1 - lon2) * math.cos(math.radians(lat2))
        v1y = (lat1 - lat2)
        v2x = (lon3 - lon2) * math.cos(math.radians(lat2))
        v2y = (lat3 - lat2)
        dot = v1x * v2x + v1y * v2y
        mag1 = math.hypot(v1x, v1y)
        mag2 = math.hypot(v2x, v2y)
        if mag1 > 0 and mag2 > 0:
            cosArg = dot / (mag1 * mag2)
            cosArg = max(-1.0, min(1.0, cosArg))
            return math.degrees(math.acos(cosArg))
    except Exception:
        pass
    return None

# Leer thresholds desde la hoja Controls si existe, si no usar defaults
def leer_thresholds(wb):
    defaults = {
        "C3": 30.0,
        "C4": 60.0,
        "C5": 120.0,
        "C6": 300.0,
        "C7": 1000.0,
        "C9": 45.0,
        "C10": 500.0,
        "C11": 800.0,
        "C15": 200.0,
        "C16": 10
    }
    if "Controls" not in wb.sheetnames:
        return defaults
    sh = wb["Controls"]
    out = {}
    for k,v in defaults.items():
        try:
            cell = sh[k].value
            if cell is None:
                out[k] = v
            else:
                if isinstance(cell,(int,float)):
                    out[k] = float(cell)
                else:
                    try:
                        out[k] = float(str(cell).strip())
                    except:
                        out[k] = v
        except:
            out[k] = v
    return out

# Mapea tramos igual que MapearTramos en VBA
def mapear_tramos(ws, start_row, last_row):
    tramoList = []
    i = start_row
    while i <= last_row:
        tipoActual = (ws.cell(row=i, column=7).value or "")
        inicio = i
        suma = 0.0
        while i <= last_row and (ws.cell(row=i, column=7).value or "") == tipoActual:
            val = ws.cell(row=i, column=5).value
            if isinstance(val, (int,float)):
                suma += float(val)
            i += 1
        tramoList.append([inicio, i-1, tipoActual, suma])
    return tramoList

def sumar_tramo(ws, pos, sentido, nivelActual, start_row, last_row):
    suma = 0.0
    while pos >= start_row and pos <= last_row and NivelKV(ws.cell(row=pos, column=7).value or "") <= nivelActual:
        val = ws.cell(row=pos, column=5).value
        if isinstance(val, (int,float)):
            suma += float(val)
        pos = pos + sentido
    return suma

def absorber(ws, pos, sentido, nivelActual, acum, tipo, start_row, last_row, threshold_c15):
    while pos >= start_row and pos <= last_row and acum < threshold_c15 and NivelKV(ws.cell(row=pos, column=7).value or "") <= nivelActual:
        val = ws.cell(row=pos, column=5).value
        if isinstance(val, (int,float)):
            acum += float(val)
        ws.cell(row=pos, column=7).value = tipo
        pos = pos + sentido
    return acum

# Colores (matching VBA RGB)
FILL_KV = {
    "KV100": PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid"),
    "KV200": PatternFill(start_color="BDECB6", end_color="BDECB6", fill_type="solid"),
    "KV300": PatternFill(start_color="51D1F6", end_color="51D1F6", fill_type="solid"),
    "KV500": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "KV1000": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
}

# --------------------
# Core: procesar_rutas
# --------------------
def procesar_rutas(rutas):
    """
    rutas: lista de objetos tal como las envia n8n (cada elemento con keys: branch, inicio, puntos)
    Devuelve lista de dicts con informacion sobre los archivos generados.
    """
    # expiration check (igual que el macro si aplica)
    if date.today() > date(2026, 6, 1):
        raise RuntimeError("Este script ha expirado (fecha límite 2026-06-01). Contacte al administrador para actualizarlo.")

    if not os.path.exists(PLANTILLA):
        raise FileNotFoundError(f"No se encuentra la plantilla: {PLANTILLA}")

    if not isinstance(rutas, list):
        raise ValueError("El JSON debe ser una lista de rutas")

    resultados = []
    print(f"Procesando {len(rutas)} rutas...")

    for rindex, ruta in enumerate(rutas, start=1):
        branch = ruta.get("branch","BR")
        inicio_name = ruta.get("inicio","START")
        puntos = ruta.get("puntos", [])
        filename = secure_filename(f"{OUTPUT_PREFIX}_{branch}_{inicio_name}.xlsx")
        salida = os.path.join(OUTPUT_DIR, filename)

        # copiar plantilla para no modificarla
        shutil.copy(PLANTILLA, salida)
        wb = load_workbook(salida)
        if "Fiber design" not in wb.sheetnames:
            print(f"ERROR: la plantilla no contiene hoja 'Fiber design' - {salida}")
            wb.close()
            resultados.append({"route_index": rindex, "error": "Hoja 'Fiber design' no encontrada"})
            continue
        ws = wb["Fiber design"]

        # Leer thresholds desde Controls
        thresholds = leer_thresholds(wb)
        C3 = thresholds["C3"]; C4 = thresholds["C4"]; C5 = thresholds["C5"]
        C6 = thresholds["C6"]; C7 = thresholds["C7"]; C9 = thresholds["C9"]
        C10 = thresholds["C10"]; C11 = thresholds["C11"]; C15 = thresholds["C15"]; C16 = int(thresholds["C16"])

        # PASO 0: limpiar E:J desde fila START_ROW hacia abajo, y A1 = "NOMBRE RUTA"
        start_row = START_ROW

        max_rows_to_clear = 20000
        for rr in range(start_row, start_row + max_rows_to_clear):
            for col in range(5, 11):  # E..J
                ws.cell(row=rr, column=col).value = None
                try:
                    ws.cell(row=rr, column=col).fill = PatternFill(fill_type=None)
                except:
                    pass

        ws["A1"] = "NOMBRE RUTA"

        # Insert points
        row = start_row
        for p in puntos:
            nombre = p.get("nombre","")
            lat = p.get("lat", None)
            lon = p.get("lon", None)
            ws.cell(row=row, column=1).value = nombre
            ws.cell(row=row, column=2).value = lat
            ws.cell(row=row, column=3).value = lon
            row += 1

        lastRow = row - 1
        if lastRow < start_row:
            print(f"Ruta {branch} {inicio_name}: sin puntos, saltando.")
            wb.save(salida); wb.close()
            resultados.append({"route_index": rindex, "skipped": True, "file": filename})
            continue

        # PASO 1: ángulos (D) y distancias (E)
        for i in range(start_row, lastRow+1):
            # ANGULO
            if i == start_row or i == lastRow:
                ws.cell(row=i, column=4).value = "N/A"
            else:
                lat1 = ws.cell(row=i-1, column=2).value
                lon1 = ws.cell(row=i-1, column=3).value
                lat2 = ws.cell(row=i, column=2).value
                lon2 = ws.cell(row=i, column=3).value
                lat3 = ws.cell(row=i+1, column=2).value
                lon3 = ws.cell(row=i+1, column=3).value
                try:
                    ang = angle_deg(lat1, lon1, lat2, lon2, lat3, lon3)
                    ws.cell(row=i, column=4).value = round(ang,2) if ang is not None else "N/A"
                except Exception:
                    ws.cell(row=i, column=4).value = "N/A"
            # DISTANCIA a siguiente punto
            if i == lastRow:
                ws.cell(row=i, column=5).value = "N/A"
            else:
                lat1 = ws.cell(row=i, column=2).value
                lon1 = ws.cell(row=i, column=3).value
                lat2 = ws.cell(row=i+1, column=2).value
                lon2 = ws.cell(row=i+1, column=3).value
                try:
                    d = distancia_m(lat1, lon1, lat2, lon2)
                    ws.cell(row=i, column=5).value = round(d,2) if d is not None else "N/A"
                except Exception:
                    ws.cell(row=i, column=5).value = "N/A"

        # PASO 2: clasificación inicial KV (col G)
        for i in range(start_row, lastRow+1):
            val = ws.cell(row=i, column=5).value
            if val is None or (isinstance(val,str) and val=="N/A"):
                ws.cell(row=i, column=7).value = ""
            else:
                try:
                    v = float(val)
                except:
                    ws.cell(row=i, column=7).value = ""
                    continue
                if v < C3:
                    ws.cell(row=i, column=7).value = "KV100"
                elif v < C4:
                    ws.cell(row=i, column=7).value = "KV200"
                elif v < C5:
                    ws.cell(row=i, column=7).value = "KV300"
                elif v < C6:
                    ws.cell(row=i, column=7).value = "KV500"
                elif v < C7:
                    ws.cell(row=i, column=7).value = "KV1000"
                else:
                    ws.cell(row=i, column=7).value = "ERROR"

        error_count = sum(1 for i in range(start_row, lastRow+1) if (ws.cell(row=i, column=7).value or "") == "ERROR")
        if error_count > 0:
            print(f"Ruta {branch} {inicio_name}: se detectaron {error_count} ERROR(s) en clasificación inicial (col G). Revisa límites en Controls.")

        # PASO 3 y 4: Mapear tramos + bucle de corrección (prioridad 5 -> 1)
        cambios = True
        iter_count = 0
        while cambios and iter_count < 200:
            iter_count += 1
            cambios = False
            tramoList = mapear_tramos(ws, start_row, lastRow)
            for lvl in range(5, 0, -1):
                for tramo in list(tramoList):
                    ini, fin, tipo, sumaDist = tramo
                    nivelActual = NivelKV(tipo)
                    if nivelActual != lvl:
                        continue
                    if sumaDist >= C15 or nivelActual >= 4:
                        continue
                    nivelArriba = NivelKV(ws.cell(row=ini-1, column=7).value or "") if ini > start_row else 0
                    nivelAbajo = NivelKV(ws.cell(row=fin+1, column=7).value or "") if fin < lastRow else 0
                    if (nivelArriba < nivelActual and nivelAbajo < nivelActual):
                        sumaArriba = sumar_tramo(ws, ini-1, -1, nivelActual, start_row, lastRow)
                        sumaAbajo = sumar_tramo(ws, fin+1, 1, nivelActual, start_row, lastRow)
                        if sumaArriba < sumaAbajo:
                            primeroArriba = True
                        elif sumaArriba > sumaAbajo:
                            primeroArriba = False
                        else:
                            if abs(nivelArriba - nivelActual) <= abs(nivelAbajo - nivelActual):
                                primeroArriba = True
                            else:
                                primeroArriba = False
                        acum = sumaDist
                        if primeroArriba:
                            acum = absorber(ws, ini-1, -1, nivelActual, acum, tipo, start_row, lastRow, C15)
                            if acum < C15:
                                acum = absorber(ws, fin+1, 1, nivelActual, acum, tipo, start_row, lastRow, C15)
                        else:
                            acum = absorber(ws, fin+1, 1, nivelActual, acum, tipo, start_row, lastRow, C15)
                            if acum < C15:
                                acum = absorber(ws, ini-1, -1, nivelActual, acum, tipo, start_row, lastRow, C15)
                        cambios = True
                        tramoList = mapear_tramos(ws, start_row, lastRow)
                        continue
                    elif (nivelArriba > nivelActual or nivelAbajo > nivelActual):
                        candUp = nivelArriba if nivelArriba > nivelActual else 0
                        candDown = nivelAbajo if nivelAbajo > nivelActual else 0
                        if candUp == 0 and candDown == 0:
                            continue
                        elif candUp == 0:
                            vecinoSuperior = candDown
                        elif candDown == 0:
                            vecinoSuperior = candUp
                        else:
                            if abs(candUp - nivelActual) < abs(candDown - nivelActual):
                                vecinoSuperior = candUp
                            elif abs(candUp - nivelActual) > abs(candDown - nivelActual):
                                vecinoSuperior = candDown
                            else:
                                vecinoSuperior = max(candUp, candDown)
                        for k in range(ini, fin+1):
                            ws.cell(row=k, column=7).value = TipoDesdeNivel(vecinoSuperior)
                        cambios = True
                        tramoList = mapear_tramos(ws, start_row, lastRow)
                        continue

        # PASO 5: pintar colores
        for i in range(start_row, lastRow+1):
            v = (ws.cell(row=i, column=7).value or "")
            if v in FILL_KV:
                try:
                    ws.cell(row=i, column=7).fill = FILL_KV[v]
                except:
                    pass

        # PASO 6: MUFA en columna H
        for i in range(start_row+1, lastRow):
            if (ws.cell(row=i, column=7).value or "") != (ws.cell(row=i-1, column=7).value or ""):
                ws.cell(row=i, column=8).value = "Mx"

        # Revision acumulada para corte cada 2860m
        acumulado = 0.0
        for i in range(start_row, lastRow):
            if (ws.cell(row=i, column=8).value or "") == "Mx":
                val = ws.cell(row=i, column=5).value
                acumulado = float(val) if isinstance(val,(int,float)) else 0.0
            else:
                v = ws.cell(row=i, column=5).value
                vnum = float(v) if isinstance(v,(int,float)) else 0.0
                if acumulado + vnum < 2860:
                    acumulado = acumulado + vnum
                else:
                    ws.cell(row=i, column=8).value = "Mx"
                    acumulado = vnum

        # PASO 7: Tension/Suspension nueva lógica para columna F (6) e I (9)
        sumaDistancia = 0.0
        cuentaSuspensiones = 0
        for i in range(start_row, lastRow+1):
            tipoKV = (ws.cell(row=i, column=7).value or "")
            if i == start_row or tipoKV != (ws.cell(row=i-1, column=7).value or ""):
                ws.cell(row=i, column=6).value = "T"
                ws.cell(row=i, column=9).value = "Inicio"
                sumaDistancia = 0.0
                cuentaSuspensiones = 0
            elif (ws.cell(row=i, column=8).value or "") == "Mx":
                ws.cell(row=i, column=6).value = "T"
                ws.cell(row=i, column=9).value = "MUFA"
                sumaDistancia = 0.0
                cuentaSuspensiones = 0
            else:
                d_ang = ws.cell(row=i, column=4).value
                ang_is_small = False
                try:
                    if isinstance(d_ang,(int,float)) and float(d_ang) < C9:
                        ang_is_small = True
                except:
                    ang_is_small = False

                if ang_is_small:
                    ws.cell(row=i, column=6).value = "T"
                    ws.cell(row=i, column=9).value = "Supera máximo angulo"
                    sumaDistancia = 0.0
                    cuentaSuspensiones = 0
                elif tipoKV in ("KV300","KV500","KV1000"):
                    ws.cell(row=i, column=6).value = "T"
                    ws.cell(row=i, column=9).value = "Tramo KV alto"
                    sumaDistancia = 0.0
                    cuentaSuspensiones = 0
                else:
                    if tipoKV == "KV100":
                        limiteDistancia = C10
                    elif tipoKV == "KV200":
                        limiteDistancia = C11
                    else:
                        limiteDistancia = 0
                    v = ws.cell(row=i, column=5).value
                    vnum = float(v) if isinstance(v,(int,float)) else 0.0
                    sumaDistancia = sumaDistancia + vnum
                    cuentaSuspensiones = cuentaSuspensiones + 1
                    if sumaDistancia > limiteDistancia or cuentaSuspensiones > C16:
                        ws.cell(row=i, column=6).value = "T"
                        ws.cell(row=i, column=9).value = "Límite superado"
                        sumaDistancia = 0.0
                        cuentaSuspensiones = 0
                    else:
                        ws.cell(row=i, column=6).value = "S"
                        ws.cell(row=i, column=9).value = ""

        # asignar "T" y motivo en última fila
        ws.cell(row=lastRow, column=6).value = "T"
        ws.cell(row=lastRow, column=9).value = "Fin de tramo"

        # Mensaje final similar al macro (col Y19 check)
        try:
            y19 = ws.cell(row=19, column=25).value  # Y=25
            if isinstance(y19,(int,float)) and y19 <= 0.4:
                ws["Z1"].value = "CHECK CAREFULLY CALCULATION"
            else:
                ws["Z1"].value = "FINALIZADO, TRA.DEPT."
        except:
            ws["Z1"].value = "FINALIZADO, TRA.DEPT."

        wb.save(salida)
        wb.close()
        print(f"✔ Ruta {rindex}/{len(rutas)} -> {salida}")
        resultados.append({"route_index": rindex, "file": filename, "path": salida})

    print("✅ Todas las rutas procesadas.")
    return resultados

# -------------------
# Flask API
# -------------------
app = Flask(__name__)

def _parse_json_from_uploaded_file(file_storage):
    raw = file_storage.stream.read(MAX_UPLOAD_BYTES + 1)
    if len(raw) > MAX_UPLOAD_BYTES:
        raise ValueError("Archivo demasiado grande")
    text = raw.decode('utf-8')
    return json.loads(text)

def _extract_rutas_from_payload(data):
    # Caso n8n full payload con items
    if isinstance(data, dict) and "items" in data and isinstance(data["items"], list):
        for it in data["items"]:
            try:
                j = it.get("json", {})
                if "rutas" in j and isinstance(j["rutas"], list):
                    return j["rutas"]
            except Exception:
                continue
        for it in data["items"]:
            try:
                j = it.get("json", {})
                if isinstance(j, list):
                    return j
            except Exception:
                continue

    if isinstance(data, dict) and "rutas" in data and isinstance(data["rutas"], list):
        return data["rutas"]

    if isinstance(data, list):
        if len(data) > 0 and isinstance(data[0], dict) and "rutas" in data[0]:
            first = data[0].get("rutas")
            if isinstance(first, list):
                return first
        if len(data) > 0 and isinstance(data[0], dict) and ("puntos" in data[0] or "branch" in data[0]):
            return data

    for key in ("body","data","payload"):
        if isinstance(data, dict) and key in data:
            return _extract_rutas_from_payload(data[key])

    return None

@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    return response

@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status":"ok"}), 200

@app.route("/procesar", methods=["POST", "OPTIONS"])
def api_procesar():
    if request.method == "OPTIONS":
        return make_response(("", 204))

    try:
        rutas = None
        data = None
        if request.files and 'file' in request.files:
            f = request.files['file']
            try:
                data = _parse_json_from_uploaded_file(f)
            except Exception as e:
                return jsonify({"error": f"Error leyendo archivo JSON: {str(e)}"}), 400
            rutas = _extract_rutas_from_payload(data)
        else:
            try:
                data = request.get_json(force=True, silent=True)
            except Exception:
                data = None

            if data is None and 'payload' in request.form:
                try:
                    data = json.loads(request.form['payload'])
                except Exception:
                    data = None

            if data is not None:
                rutas = _extract_rutas_from_payload(data)

        if rutas is None:
            return jsonify({"error":"No se pudo extraer la lista de rutas. Envíe JSON con 'rutas' o una lista de rutas."}), 400

        if not isinstance(rutas, list):
            return jsonify({"error":"El JSON debe contener una lista de rutas."}), 400

        callback_url = request.args.get("callback_url") or (data.get("callback_url") if isinstance(data, dict) else None)

        resultados = procesar_rutas(rutas)

        host = request.host_url.rstrip('/')
        for r in resultados:
            if "file" in r:
                r["download_url"] = f"{host}/download/{r['file']}"

        resp_body = {"status":"ok", "results": resultados}

        callback_status = None
        if callback_url:
            try:
                cb = requests.post(callback_url, json=resp_body, timeout=20)
                callback_status = {"status_code": cb.status_code, "ok": cb.ok}
            except Exception as e:
                callback_status = {"error": str(e)}
            resp_body["callback_status"] = callback_status

        return jsonify(resp_body), 200

    except FileNotFoundError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        print("Error en /procesar:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route("/download/<filename>", methods=["GET"])
def download_file(filename):
    safe_name = secure_filename(filename)
    file_path = os.path.join(OUTPUT_DIR, safe_name)
    if not os.path.exists(file_path):
        return jsonify({"error":"file not found"}), 404
    return send_from_directory(OUTPUT_DIR, safe_name, as_attachment=True)

if __name__ == "__main__":
    # En producción usar gunicorn: e.g. gunicorn -w 4 "calcular_rutas_full_complete:app"
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
